## -*- coding: utf-8 -*-
import shutil
#from typing import Tuple
from joern.all import JoernSteps
from igraph import *
from general_op2 import *
import pickle
import os
import time
import json
from py2neo.packages.httpstream import http
from optparse import OptionParser
import openpyxl
from multiprocessing import Process, Lock
http.socket_timeout = 9999
'''
2022.3.29Modified:from the previous only get a single vulnerability function, to get a list of vulnerability functions
'''

f = open('./config.json')
path_data = json.load(f)

def get_calls_id(db, func_name):
    query_str = 'getCallsTo("%s")' % func_name
    results = db.runGremlinQuery(query_str)
    return results

def getNodesByID(db, func_id):
    query_str = 'queryNodeIndex("type:FunctionDef AND functionId:%s")' % func_id
    node = db.runGremlinQuery(query_str)
    
    return node

def getFuncNode(db, func_name):
    query_str = 'getFunctionsByName("' + func_name + '")'
    func_node = db.runGremlinQuery(query_str)
    print(func_node)
    return func_node

def getFuncName(code):
    pattern = "((?:_|[A-Za-z])\w*(?:\s(?:\.|::|\->|)\s(?:_|[A-Za-z])\w*)*)\s\("
    result = re.findall(pattern, code)
    return result

def getFuncFile(db, func_id):
    #print(type(db))
    query_str = "g.v(%d).in('IS_FILE_OF').filepath" % func_id
    ret = db.runGremlinQuery(query_str)
    #print ret
    return ret[0]

def getCalleeNode(db, func_id):
    query_str = "queryNodeIndex('type:Callee AND functionId:%d')" % func_id
    results = db.runGremlinQuery(query_str)
    return results


def get_all_calls_node(db, func_name):
    list_all_funcID = [node._id for node in getFuncNode(db, func_name)]
    #print "list_all_funcID", list_all_funcID
    #print "lenth", len(list_all_funcID)
    if len(list_all_funcID)>1030:
        print ">1000"
        return False
    list_all_callee_node = []
    for func_id in list_all_funcID:#allfile in a testID
        list_all_callee_node += getCalleeNode(db, func_id)

    if list_all_callee_node == []:
        return False
    else:
        #return [node for node in list_all_callee_node]
        return [(str(node._id), node.properties['code'], str(node.properties['functionId'])) for node in list_all_callee_node]

#The function that reads the vulnerability corresponding to each CVE returns the corresponding dictionary
def get_vulname(software):
    dst_folder = './data/C-Diffs/' + software #diff
    # xlsx_path = './' + software + '_data.xlsx'
    xlsx_path = './all_data.xlsx'
    ws = openpyxl.load_workbook(xlsx_path)['Sheet1']
    #ws = openpyxl.load_workbook(xlsx_path)['ffmpeg']
    dic = {}
    vulres = []

    rows = ws.max_row
    columns = ws.max_column

    for i in range(2, rows):
        cve = ws.cell(row=i, column=1).value
        hashs = ws.cell(row=i, column=2).value
        vulname = ws.cell(row=i, column=4).value
        if(vulname == ''):
            continue

        dic.setdefault(cve+hashs, []).append(vulname)
    
    return dic

def delete_dir(fold):
    shutil.rmtree(fold)
    os.mkdir(fold)

if __name__ == '__main__':
    
    parser = OptionParser()
    (options, args) = parser.parse_args()
    if(len(args) != 1):
        print('Missing parameters! Please add software name.')
    
    #Delete the last generated key variable file and intermediate output file
    os.chdir('./cv_result')
    os.system('rm *')
    print('delete * in ./cv_result')
    os.chdir(path_data["start_folder"])

    start_dic = get_vulname(args[0])
    #start = ['ebml_parse_elem']
    os.chdir(path_data["start_folder"]) #Return to the src directory
    os.system('python3 cv_extract.py')
    print('get cv_extract.py')
    print(start_dic)
    dst_folder = './data/C-Diffs/' + args[0]
    #print(dst_folder)
    #lock = Lock()
    for cve in os.listdir(dst_folder):
        diff_path = dst_folder + '/' + cve
        print(diff_path)
        for diff in os.listdir(diff_path):
            hashs = diff.split('_')[2]
            
            index = cve+hashs
            print(index)
            if(index not in start_dic.keys()):
                continue
            start = start_dic[index]

            
            #print("joern is over")
            os.chdir(path_data["joern"]["joern_exe"])
            #print(os.access('.joernIndex', os.W_OK))
            if(os.path.exists('.joernIndex')):
                os.system('chmod +777 .joernIndex')
                shutil.rmtree(path_data["joern"]["before_joern"]) #Delete the previous legacy .joernIndex file
            print('delete .joernIndex')
            os.system('ls -a')
            src = path_data["start_folder"] + './gitrepos/' + args[0] + '_git'
            dst = path_data["joern"]["testCode"]
            os.chdir(src)
            os.system('git stash')
            
            commands = 'git checkout {0}'.format(hashs)
            #os.system('git stash')
            os.system(commands)

            #os.remove(dst)
            delete_dir(dst) #Delete the files to be parsed from the last joern/testCode
            

            os.chdir(path_data["neo4j"])
            os.system("./neo4j stop")
            os.system("./neo4j status")
            cnt = 0
            for root, dirs, files in os.walk(src):
                for f in files:
                    if(f == 'parse_date.c'):
                        continue
                    if('.c' in f):
                        s = os.path.join(root, f)
                        shutil.copy(s, dst)
                        cnt += 1

            print(str(cnt) + 'files has been copy.')

            os.chdir(path_data["joern"]["joern_exe"])
            os.system('./joern testCode')
            print('joern is over')

            os.system('../neo4j/bin/./neo4j start-no-wait')
            os.system('sleep 15s')

            j = JoernSteps()
            j.connectToDatabase()

            filename1 = []

            start1 = start
            cnt = 3

            while(cnt > 0):
                tmp = []
                print(type(start1))
                print(start1)
                #os.system('./neo4j status')
                for name in start1:
                    print(type(name))
                    results = get_all_calls_node(j, str(name))
                    print(results)
                    
                    if(results == False):
                        continue
                    for i in results:
                        #print(i[0], i[1], i[2])
                        funcname = i[1]
                        print(funcname)
                        funcnode = getFuncNode(j, funcname)
                        #funcnode = j.runGremlinQuery('getFunctionsByName(' + funcname + ')')
                        if(funcnode == []):
                            print('not found funcnode.')
                            continue

                        filename1.append(getFuncFile(j, funcnode[0]._id))
                        if(i[1] not in tmp):
                            print(type(i[1]))
                            tmp.append(str(i[1]))
                            print(i[1], getFuncFile(j, int(i[2])))
                        #print(i[1])
                start1 = tmp
                cnt -= 1
                print('-------------------------------------------------')

            cnt = 3
            filename2 = []
            '''This is the case for obtaining only a single vulnerability function
            start2 = []
            start2.append(str(start))
            '''
            start2 = start
            while(cnt > 0):
                print(start2)
                tmp = []
                for name in start2:
                    results = get_calls_id(j, name)
                    #results = j.runGremlinQuery('getCallsTo(' + name + ')')
                    print(results)
                    if(results == []):
                        continue

                    for i in results:
                        funcid = i.properties['functionId']
                        filename2.append(getFuncFile(j, funcid))
                        calls_node = getNodesByID(j, funcid)[0]
                        res = getFuncName(calls_node.properties['code'])
                        print(funcid, res)
                        for filen in res:
                            if(filen not in tmp):
                                tmp.append(str(filen))
                print('-------------------------------')
                start2 = tmp
                cnt -= 1

            file_tmp = set(filename2 + filename1)
            file_res = [i for i in file_tmp]
            os.chdir(path_data['start_folder'])
            #depen_txt = path_data["all_test_code"]["all_dep_path"] + cve + '/depenfile_' + hashs + '.txt'
            #depen_txt = './depenfile.txt'
            depen_path = path_data["all_test_code"]["all_dep_path"] + args[0] + '/' + cve

            if(not os.path.exists(depen_path)):
                os.mkdir(depen_path)
            depen_txt = depen_path + '/depenfile_' + hashs + '.txt'
            if(os.path.exists(depen_txt)):
                continue
            f = open(depen_txt, 'wb+')

            for i in file_res:
                f.write(i)
                f.write('\n')
            f.close()
            


